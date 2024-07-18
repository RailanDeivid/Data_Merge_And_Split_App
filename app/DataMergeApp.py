import streamlit as st
import pandas as pd
import tempfile
import io
import os
from streamlit_option_menu import option_menu
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook



# Configuração da página
st.set_page_config(page_title="DataMergeApp",
                   page_icon=":file_folder:", layout="wide")
st.markdown("""
    <style>
        .contact-me {
            text-align: center; /* Centraliza o texto */
            font-size: 12px;
            color: hsl(0, 0%, 35%);
            
        }
        .linkedin-icon {
            width: 15px;
            vertical-align: middle;
            margin-left: 5px;
        }
    </style>
    """, unsafe_allow_html=True)

st.markdown("<p class='contact-me'>💡 Desenvolvido por Railan Deivid<br>Contate-me <a href='https://www.linkedin.com/in/railandeivid/' target='_blank'><img src='https://cdn-icons-png.flaticon.com/512/174/174857.png' class='linkedin-icon'></a></p>", unsafe_allow_html=True)

# ----------------------------------------------- Função para combinar arquivos ------------------------------------- #
def combinar_arquivos():
    # Setar título
    st.markdown("""
        <style>
            .rounded-title {
                text-align: center; 
                font-size: 40px;
            }
        </style>
        """, unsafe_allow_html=True)

    st.markdown("<h1 class='rounded-title'>Combinar Arquivos</h1><br>", unsafe_allow_html=True)
    st.write("Faça upload de múltiplos arquivos XLSX ou CSV para combiná-los em um único arquivo XLSX.")
    st.info("""Certifique-se de que os arquivos tenham a mesma quantidade de colunas com as mesmas nomenclaturas.
            Caso sejam arquivos de Excel com várias abas a serem combinadas, verifique se os nomes das abas são consistentes em todos os arquivos selecionados.""", icon=":material/info:")

    # Upload dos arquivos
    uploaded_files = st.file_uploader("Escolha arquivos XLSX ou CSV", accept_multiple_files=True, type=['xlsx', 'csv'])

    if uploaded_files:
        # Limpa o st.session_state se um novo arquivo for carregado
        if 'last_uploaded_files' in st.session_state and st.session_state.last_uploaded_files != [file.name for file in uploaded_files]:
            st.session_state.filtered_data_dict = {}
            st.session_state.separated_data = False
        st.session_state.last_uploaded_files = [file.name for file in uploaded_files]

        # Checkbox para inserir ou não a coluna de origem, aparece apenas se o arquivo for XLSX
        if any(file.name.endswith('.xlsx') for file in uploaded_files):
            add_origin_column = st.checkbox("Adicionar coluna com o nome da origem", value=False)
            origin_column_name = None
            create_column = False

            if add_origin_column:
                origin_column_name = st.text_input("Nome da coluna de origem", value="Aba Origem")
                create_column = st.button("Criar Coluna")

        # Opções para separador e encoding
        if any(file.name.endswith('.csv') for file in uploaded_files):
            cols = st.columns(5)
            with cols[0]:
                sep = st.selectbox("Selecione o separador CSV:", [',', ';'])
            with cols[1]:
                encoding = st.selectbox("Selecione o encoding do arquivo CSV:", ['utf-8', 'latin1', 'iso-8859-1'])

        # Lista para armazenar DataFrames de cada arquivo
        dfs = []

        for file in uploaded_files:
            # Verifica se o arquivo é XLSX ou CSV
            if file.name.endswith('.xlsx'):
                # Carrega o arquivo Excel
                xls = pd.ExcelFile(file)

                # Lista de abas disponíveis no arquivo
                sheet_names = xls.sheet_names

                # Checkbox para seleção de todas as abas (inicializado como True por padrão)
                select_all_sheets = st.checkbox(f"Selecionar todas as abas de '{file.name}'", value=True)

                if select_all_sheets:
                    selected_sheets = sheet_names
                else:
                    selected_sheets = st.multiselect(f"Selecione as abas de '{file.name}'", sheet_names, default=sheet_names)

                for sheet in selected_sheets:
                    # Lê o DataFrame da aba selecionada
                    df = pd.read_excel(file, sheet_name=sheet)

                    # Adiciona uma coluna com o nome da aba origem se a opção estiver selecionada e o botão for clicado
                    if add_origin_column and create_column:
                        df[origin_column_name] = sheet

                    dfs.append(df)

            elif file.name.endswith('.csv'):
                try:
                    # Lê o DataFrame do arquivo CSV com as opções selecionadas
                    df = pd.read_csv(file, sep=sep, encoding=encoding)
                    nome_sheet = file.name.rfind('.')
                    df['Arquivo Origem'] = file.name[:nome_sheet]
                    dfs.append(df)
                except UnicodeDecodeError:
                    st.error(f"Erro ao ler o arquivo '{file.name}'. Verifique o encoding selecionado.")

        if dfs:
            # Concatena todos os DataFrames em um único DataFrame
            combined_data = pd.concat(dfs, ignore_index=True)

            # Mostra os dados combinados
            st.subheader("Dados Combinados:")
            st.write(combined_data.head(10))

            # Download do arquivo combinado
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                combined_data.to_excel(writer, index=False, sheet_name='DadosCombinados')

            # Reabertura do arquivo com openpyxl para adicionar a formatação da tabela
            output.seek(0)
            workbook = load_workbook(output)
            worksheet = workbook['DadosCombinados']
            table = Table(displayName='DadosCombinadosTable', ref=worksheet.dimensions)
            style = TableStyleInfo(name='TableStyleLight9', showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            worksheet.add_table(table)

            # Salvando novamente o arquivo formatado
            formatted_output = io.BytesIO()
            workbook.save(formatted_output)
            formatted_output.seek(0)

            # Botão de download com o nome personalizado do arquivo
            st.download_button(
                label="Baixar Dados Combinados",
                data=formatted_output,
                file_name="DadosCombinados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Botão para limpar os dados armazenados em st.session_state
            st.markdown("<br>", unsafe_allow_html=True)
            st.info("Antes de fazer outra operação, limpe os dados!",
                    icon=":material/warning:")
            if st.button("Limpar Dados"):
                st.session_state.filtered_data_dict = {}
                st.session_state.separated_data = False
                st.experimental_rerun()
    pass



def separar_arquivos():
    # Setar título
    st.markdown("""
        <style>
            .rounded-title {
                text-align: center; 
                font-size: 40px;
            }
        </style>
        """, unsafe_allow_html=True)

    st.markdown("<h1 class='rounded-title'>Separar Arquivos</h1><br>",
                unsafe_allow_html=True)

    st.write("Selecione um arquivo XLSX para separá-lo com base em uma coluna ou aba específica ou um arquivo CSV")

    # Upload do arquivo
    uploaded_file = st.file_uploader(
        "Escolha um arquivo XLSX ou CSV", type=['xlsx', 'csv'])

    if uploaded_file:
        # Limpa o st.session_state se um novo arquivo for carregado
        if 'last_uploaded_file' in st.session_state and st.session_state.last_uploaded_file != uploaded_file.name:
            st.session_state.filtered_data_dict = {}
            st.session_state.separated_data = False
        st.session_state.last_uploaded_file = uploaded_file.name

        if uploaded_file.name.endswith('.xlsx'):
            # Se o arquivo for XLSX, mostra opções para selecionar aba e coluna
            xls = pd.ExcelFile(uploaded_file)
            sheet_names = xls.sheet_names

            cols = st.columns(3)
            with cols[0]:
                selected_sheet = st.selectbox("Selecione a aba:", sheet_names)

            # Lê o DataFrame da aba selecionada
            df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)

            # Seleção da coluna para separação
            col_options = df.columns.tolist()
            with cols[1]:
                selected_column_1 = st.selectbox("Selecione a primeira coluna para separar os dados:", col_options)
            with cols[2]:
                metodo_separacao = st.radio("Escolha o método de separação:", ('Separar por uma coluna', 'Separar por coluna e abas'))
                if metodo_separacao == 'Separar por coluna e abas':
                    selected_column_2 = st.selectbox("Selecione a segunda coluna para separar os dados:", col_options)

        elif uploaded_file.name.endswith('.csv'):
            try:
                # Se o arquivo for CSV, mostra opções para selecionar separador, encoding e coluna
                cols = st.columns(4)
                with cols[0]:
                    sep = st.selectbox("Selecione o separador CSV:", [',', ';'])
                with cols[1]:
                    encoding = st.selectbox("Selecione o encoding do arquivo CSV:", ['utf-8', 'latin1', 'iso-8859-1'])

                # Lê o DataFrame do arquivo CSV com as opções selecionadas
                df = pd.read_csv(uploaded_file, sep=sep, encoding=encoding)

                # Seleção da coluna para separação
                col_options = df.columns.tolist()
                with cols[2]:
                    selected_column_1 = st.selectbox("Selecione a coluna para separar os dados:", col_options)

            except UnicodeDecodeError:
                st.error(f"Erro ao ler o arquivo '{uploaded_file.name}'. Verifique o encoding selecionado.")
                return

        # Mostra os dados originais
        st.subheader("Dados Originais:")
        st.write(df.head(10))

        # Inicializa st.session_state se necessário
        if 'filtered_data_dict' not in st.session_state:
            st.session_state.filtered_data_dict = {}
        if 'separated_data' not in st.session_state:
            st.session_state.separated_data = False

        # Cria um botão para separar os dados com base na escolha do usuário
        if st.button("Separar Dados"):
            # Para cada valor único na primeira coluna selecionada
            for valor_coluna_1 in df[selected_column_1].unique():
                # Filtra o DataFrame com base no valor único da primeira coluna
                dados_filtrados_1 = df[df[selected_column_1] == valor_coluna_1]

                # Inicializa um novo arquivo Excel para cada valor único da primeira coluna
                nome_arquivo = f'{valor_coluna_1}.xlsx'
                wb = Workbook()
                default_sheet = wb.active
                wb.remove(default_sheet)

                if uploaded_file.name.endswith('.xlsx') and metodo_separacao == 'Separar por coluna e abas':
                    # Para cada valor único na segunda coluna dentro do DataFrame filtrado pela primeira coluna
                    for valor_coluna_2 in dados_filtrados_1[selected_column_2].unique():
                        # Filtra o DataFrame com base no valor único da segunda coluna
                        dados_filtrados_2 = dados_filtrados_1[dados_filtrados_1[selected_column_2]
                                                              == valor_coluna_2]

                        # Limpar espaços em branco do nome da aba
                        nome_aba = str(valor_coluna_2).strip()

                        # Verifique se o nome da aba é válido (não pode conter espaços)
                        if isinstance(nome_aba, str) and ' ' in nome_aba:
                            nome_aba = nome_aba.replace(' ', '_')

                        # Criar a aba no workbook
                        ws = wb.create_sheet(title=nome_aba)

                        # Adiciona os dados do DataFrame filtrado à planilha
                        df_filtrado_2 = pd.DataFrame(dados_filtrados_2)
                        df_filtrado_2 = df_filtrado_2.sort_values(
                            by=[selected_column_1, selected_column_2], ascending=True)

                        for row in dataframe_to_rows(df_filtrado_2, index=False, header=True):
                            ws.append(row)

                        # Formata a área dos dados como uma tabela
                        tab = Table(displayName=str(
                            ws.title), ref=ws.dimensions)
                        tab.tableStyleInfo = TableStyleInfo(name='TableStyleLight9', showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                        ws.add_table(tab)

                else:
                    # Se o método de separação for apenas por uma coluna
                    # Verifique se o nome da aba é válido (não pode conter espaços)
                    if isinstance(valor_coluna_1, str) and ' ' in valor_coluna_1:
                        valor_coluna_1 = valor_coluna_1.replace(' ', '_')

                    ws = wb.create_sheet(title=str(valor_coluna_1))

                    # Adiciona os dados do DataFrame filtrado à planilha
                    df_filtrado_1 = pd.DataFrame(dados_filtrados_1)
                    df_filtrado_1 = df_filtrado_1.sort_values(
                        by=[selected_column_1], ascending=True)

                    for row in dataframe_to_rows(df_filtrado_1, index=False, header=True):
                        ws.append(row)

                    # Formata a área dos dados como uma tabela
                    tab = Table(displayName=str(ws.title), ref=ws.dimensions)
                    tab.tableStyleInfo = TableStyleInfo(name='TableStyleLight9', showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                    ws.add_table(tab)

                # Guarda os dados filtrados em st.session_state
                st.session_state.filtered_data_dict[valor_coluna_1] = nome_arquivo
                wb.save(nome_arquivo)
            st.session_state.separated_data = True

        # Exibe e disponibiliza para download os dados separados armazenados em st.session_state
        if st.session_state.separated_data:
            for valor_coluna_1, nome_arquivo in st.session_state.filtered_data_dict.items():
                st.markdown("---")
                st.subheader(f"Dados separados para '{valor_coluna_1}':")

                # Mostra os dados tratados se for XLSX com método 'Separar por uma coluna'
                if uploaded_file.name.endswith('.xlsx') and metodo_separacao == 'Separar por uma coluna':
                    df_tratado = pd.read_excel(nome_arquivo, sheet_name=str(
                        valor_coluna_1).replace(" ", "_").replace("/", "_"))
                    st.write(df_tratado.head())

                # Download do arquivo separado
                with open(nome_arquivo, 'rb') as file:
                    st.download_button(
                        label=f"Baixar Dados para '{valor_coluna_1}'",
                        data=file,
                        file_name=nome_arquivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            # Botão para limpar os dados armazenados em st.session_state
            st.markdown("<br>", unsafe_allow_html=True)
            st.info("Antes de fazer outra operação, limpe os dados!",
                    icon=":material/warning:")
            if st.button("Limpar Dados"):
                st.session_state.filtered_data_dict = {}
                st.session_state.separated_data = False
                st.experimental_rerun()


# ------------------------------------------------------ Menu de navegação usando option_menu ------------------------ #
cols1, cols2, cols3 = st.columns([1, 1.3, 1])
with cols2:
    selected_page = option_menu(
        menu_title=None,
        options=["Combinar Arquivos", "Separar Arquivos"],
        icons=["files", "file-earmark-break"],
        menu_icon="cast",
        default_index=0,
        orientation="horizontal"
    )

# Lógica de seleção da página
if selected_page == "Combinar Arquivos":
    combinar_arquivos()

elif selected_page == "Separar Arquivos":
    separar_arquivos()
